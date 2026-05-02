"""
File-Agnostic Timetable Data Parser

APPROACH:
---------
This module provides a TimetableDataParser class that converts timetable data from multiple
input formats (Excel, CSV, JSON) into a normalized JSON intermediate format compatible with
the IngestionService. The parser is format-agnostic: file type is auto-detected from extension.

The parser follows a two-phase validation strategy:

  Phase 1 (Sheet/Table-level): Parse each data source independently, collecting all errors
                               (name uniqueness, field presence, type constraints).
  Phase 2 (Cross-source): Validate foreign key-like references across sources
                          (professor/room/student_group lookups in the Courses data).

The parser reports ALL errors before exiting (not fail-fast), so users see every issue in one run.
Optional fields are always coerced to their defaults if absent or invalid. A --strict flag
can promote warnings (e.g., missing optional fields) to hard errors.

SUPPORTED FORMATS:
------------------
  - .xlsx (Excel): 4 sheets (Professors, Rooms, Student Groups, Courses)
  - .csv (CSV): Single file with course instances; auto-creates professor/room/group records
  - .json (JSON): Pre-normalized structure matching IngestionService expectations

USAGE:
------
  1. Programmatic (Python):
     >>> from schema.parser import TimetableDataParser
     >>> parser = TimetableDataParser()
     >>> success, stats = parser.parse("data.xlsx", "out.json")
     >>> if success:
     ...     print(f"Parsed {stats['counts']['courses']} courses")

  2. Command-line:
     $ python schema/parser.py --input data.xlsx --output out.json
     $ python schema/parser.py --input data.csv --output out.json --strict
     $ python schema/parser.py --input seed.json --output normalized.json

ASSUMPTIONS:
------------
  A1: Input file format is auto-detected from extension (.xlsx, .csv, .json)

  A2: For .xlsx files:
      - Exactly 4 named sheets (case-insensitive): "Professors", "Rooms", "Student Groups", "Courses"
      - First row of each sheet is a header row; column names are case-insensitive

  A3: For .csv files:
      - Single file with course/session data (one row per course instance)
      - Required columns: course_code, session_type, professor, room, student_group
      - Professors/rooms/student_groups referenced in CSV are auto-created if not present

  A4: For .json files:
      - Input can be pre-normalized (with professors/rooms/student_groups/courses keys)
      - Or a flat array of course instances (auto-expands to full structure)

  A5: Professors: 'name' required; 'email' optional (auto-generated from name)
  A6: Rooms: 'name' required; 'is_lab' (bool, default False), 'capacity' (int, default 100)
  A7: Student Groups: 'name', 'size' required; 'level' optional (default 'batch')
  A8: Courses: 'course_code', 'session_type', 'professor', 'room', 'student_group' required;
                others optional with sensible defaults
  A9: Cross-source references must match exactly (dangling refs = hard error)
  A10: Output JSON location controlled by output_path; default: schema/output/parsed_data.json

INTEGRATION:
------------
The generated JSON can be directly passed to the IngestionService:
  >>> from backend.app.services.ingestion_service import IngestionService
  >>> result = IngestionService.perform_initial_seeding("out.json")

Or via the Flask API:
  POST /api/ingestion/seed with custom seed file path

ERROR REPORTING:
----------------
All errors include: source name, row number, column name, and description.
Example error output:
  ERROR [Courses, Row 5, Column 'professor']: Unknown professor 'Dr. Unknown'
  ERROR [Rooms, Row 3, Column 'capacity']: Expected positive integer, got 'abc'
  WARNING [Courses, Row 2, Column 'total_credits']: Missing optional field; using default 3
"""

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Worksheet = None


# ============================================================================
# Constants
# ============================================================================

REQUIRED_SHEETS = {"Professors", "Rooms", "Student Groups", "Courses"}
SESSION_TYPES = {"lecture", "tutorial", "lab"}
VALID_LEVELS = {"batch", "department", "track", "elective"}
VALID_BINS = {1, 2, 3}
DEFAULT_EMAIL_DOMAIN = "@mahindrauniversity.edu.in"


# ============================================================================
# Helper Functions
# ============================================================================

def normalize_sheet_name(name: str) -> str:
    """Normalize sheet name for case-insensitive matching."""
    return name.strip().lower() if name else ""


def normalize_header(header: str) -> str:
    """Normalize column header for case-insensitive matching."""
    return header.strip().lower() if header else ""


def coerce_bool(value: Any) -> bool:
    """Coerce value to boolean. Accepts: True, 1, 'yes', 'true', 'lab', etc."""
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "1", "lab"}
    return False


def coerce_int(value: Any, allow_zero: bool = False) -> Optional[int]:
    """Coerce value to int. Returns None if invalid."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        try:
            val = int(value.strip())
            return val if (val > 0 or allow_zero) else None
        except ValueError:
            return None
    return None


def auto_generate_email(name: str) -> str:
    """Generate email from professor name (strip 'Dr.' prefix)."""
    clean = re.sub(r"^dr\.\s*", "", name, flags=re.IGNORECASE).strip()
    return f"{clean.lower().replace(' ', '.')}{DEFAULT_EMAIL_DOMAIN}"


# ============================================================================
# TimetableDataParser Class
# ============================================================================

class TimetableDataParser:
    """
    File-agnostic parser for timetable data.

    Attributes:
        strict: If True, warnings become hard errors
    """

    def __init__(self, strict: bool = False):
        """
        Initialize parser (file-agnostic, no file args required).

        Args:
            strict: If True, treat warnings as hard errors
        """
        self.strict = strict
        self._reset_state()

    def _reset_state(self) -> None:
        """Reset internal state for a new parse operation."""
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.professors: List[Dict[str, str]] = []
        self.rooms: List[Dict[str, Any]] = []
        self.student_groups: List[Dict[str, Any]] = []
        self.courses: List[Dict[str, Any]] = []
        self.professor_names: Set[str] = set()
        self.room_names: Set[str] = set()
        self.student_group_names: Set[str] = set()

    def parse(self, input_path: str, output_path: str = "schema/output/parsed_data.json") -> Tuple[bool, Dict[str, Any]]:
        """
        Parse file and generate JSON output. Auto-detects format from extension.

        Args:
            input_path: Path to input file (.xlsx, .csv, or .json)
            output_path: Path to output JSON file (default: schema/output/parsed_data.json)

        Returns:
            (success: bool, stats: dict) where stats includes:
              - 'counts': {professors, rooms, student_groups, courses}
              - 'errors': list of error messages
              - 'warnings': list of warning messages
        """
        self._reset_state()
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)

        ext = self.input_path.suffix.lower()

        if ext == ".xlsx":
            self._parse_excel()
        elif ext == ".csv":
            self._parse_csv()
        elif ext == ".json":
            self._parse_json()
        else:
            self.errors.append(f"ERROR: Unsupported file format '{ext}'. Use .xlsx, .csv, or .json")
            return False, self._build_stats()

        self._cross_validate()

        if self.errors or (self.strict and self.warnings):
            self._print_errors()
            return False, self._build_stats()

        self._write_json()
        return True, self._build_stats()

    def _parse_excel(self) -> None:
        """Parse Excel file with 4 sheets."""
        if not HAS_OPENPYXL:
            self.errors.append("ERROR: openpyxl is required for Excel parsing. Install with: pip install openpyxl")
            return

        try:
            workbook = load_workbook(self.input_path, data_only=True)
        except Exception as e:
            self.errors.append(f"ERROR: Failed to load workbook: {e}")
            return

        sheet_names = {normalize_sheet_name(s): s for s in workbook.sheetnames}

        missing = [s for s in REQUIRED_SHEETS if normalize_sheet_name(s) not in sheet_names]
        if missing:
            self.errors.append(f"ERROR: Missing required sheets: {', '.join(missing)}")
            return

        self._parse_professors(sheet_names, workbook)
        self._parse_rooms(sheet_names, workbook)
        self._parse_student_groups(sheet_names, workbook)
        self._parse_courses(sheet_names, workbook)

    def _parse_csv(self) -> None:
        """Parse CSV file with course instances. Auto-creates professors/rooms/groups."""
        try:
            import csv
            with open(self.input_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    self.errors.append("ERROR: CSV file is empty")
                    return

                for row_num, row in enumerate(reader, start=2):
                    # Auto-create referenced entities
                    if "professor" in row and row["professor"] and row["professor"] not in self.professor_names:
                        self._add_professor_from_dict({"name": row["professor"]}, source="CSV", row_num=row_num)

                    if "room" in row and row["room"] and row["room"] not in self.room_names:
                        self._add_room_from_dict({"name": row["room"]}, source="CSV", row_num=row_num)

                    if "student_group" in row and row["student_group"] and row["student_group"] not in self.student_group_names:
                        self._add_student_group_from_dict(
                            {"name": row["student_group"], "size": 80, "level": "batch"},
                            source="CSV", row_num=row_num
                        )

                    # Now add the course
                    self._add_course_from_dict(row, source="CSV", row_num=row_num)
        except Exception as e:
            self.errors.append(f"ERROR: Failed to parse CSV: {e}")

    def _parse_json(self) -> None:
        """Parse JSON file (pre-normalized or flat array of course instances)."""
        try:
            with open(self.input_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            self.errors.append(f"ERROR: Failed to parse JSON: {e}")
            return

        if isinstance(data, list):
            # Flat array: auto-create entities, then add courses
            for row_num, item in enumerate(data, start=1):
                # Auto-create referenced entities
                if "professor" in item and item["professor"] and item["professor"] not in self.professor_names:
                    self._add_professor_from_dict({"name": item["professor"]}, source="JSON", row_num=row_num)

                if "room" in item and item["room"] and item["room"] not in self.room_names:
                    self._add_room_from_dict({"name": item["room"]}, source="JSON", row_num=row_num)

                if "student_group" in item and item["student_group"] and item["student_group"] not in self.student_group_names:
                    self._add_student_group_from_dict(
                        {"name": item["student_group"], "size": 80, "level": "batch"},
                        source="JSON", row_num=row_num
                    )

                self._add_course_from_dict(item, source="JSON", row_num=row_num)

        elif isinstance(data, dict):
            # Pre-normalized: parse in order (professors, rooms, groups, then courses)
            if "professors" in data:
                for prof in data["professors"]:
                    self._add_professor_from_dict(prof, source="JSON")

            if "rooms" in data:
                for room in data["rooms"]:
                    self._add_room_from_dict(room, source="JSON")

            if "student_groups" in data:
                for group in data["student_groups"]:
                    self._add_student_group_from_dict(group, source="JSON")

            if "courses" in data:
                for row_num, course in enumerate(data["courses"], start=1):
                    # Auto-create missing entities
                    if "professor" in course and course["professor"] and course["professor"] not in self.professor_names:
                        self._add_professor_from_dict({"name": course["professor"]}, source="JSON", row_num=row_num)

                    if "room" in course and course["room"] and course["room"] not in self.room_names:
                        self._add_room_from_dict({"name": course["room"]}, source="JSON", row_num=row_num)

                    if "student_group" in course and course["student_group"] and course["student_group"] not in self.student_group_names:
                        self._add_student_group_from_dict(
                            {"name": course["student_group"], "size": 80, "level": "batch"},
                            source="JSON", row_num=row_num
                        )

                    self._add_course_from_dict(course, source="JSON", row_num=row_num)
        else:
            self.errors.append("ERROR: JSON must be an object or array")

    def _add_professor_from_dict(self, prof_dict: Dict[str, Any], source: str = "source", row_num: int = 0) -> None:
        """Add a professor from a dictionary (used by all format parsers)."""
        name = prof_dict.get("name", "").strip() if isinstance(prof_dict.get("name"), str) else ""

        if not name:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'name']: Required field is empty"
            )
            return

        if name in self.professor_names:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + f", Column 'name']: Duplicate professor name '{name}'"
            )
            return

        email = prof_dict.get("email", "").strip() if isinstance(prof_dict.get("email"), str) else ""
        if not email:
            email = auto_generate_email(name)
            self.warnings.append(
                f"WARNING [{source}" + (f", Row {row_num}" if row_num else "") + f", Column 'email']: Missing email; auto-generated: {email}"
            )

        self.professor_names.add(name)
        self.professors.append({"name": name, "email": email})

    def _parse_professors(self, sheet_names: Dict[str, str], workbook) -> None:
        """Parse Professors sheet (Excel only)."""
        sheet_key = normalize_sheet_name("Professors")
        sheet_name = sheet_names[sheet_key]
        sheet = workbook[sheet_name]

        rows = self._sheet_to_dicts(sheet, sheet_name)
        if not rows:
            self.errors.append(f"ERROR [{sheet_name}]: Sheet is empty")
            return

        for row_num, row_dict in enumerate(rows, start=2):
            self._add_professor_from_dict(row_dict, source=sheet_name, row_num=row_num)

    def _add_room_from_dict(self, room_dict: Dict[str, Any], source: str = "source", row_num: int = 0) -> None:
        """Add a room from a dictionary (used by all format parsers)."""
        name = room_dict.get("name", "").strip() if isinstance(room_dict.get("name"), str) else ""

        if not name:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'name']: Required field is empty"
            )
            return

        if name in self.room_names:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + f", Column 'name']: Duplicate room name '{name}'"
            )
            return

        is_lab = coerce_bool(room_dict.get("is_lab"))
        capacity = coerce_int(room_dict.get("capacity")) or 100

        if capacity <= 0:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'capacity']: "
                f"Expected positive integer, got {room_dict.get('capacity')}"
            )
            return

        self.room_names.add(name)
        self.rooms.append({
            "name": name,
            "is_lab": is_lab,
            "capacity": capacity,
        })

    def _parse_rooms(self, sheet_names: Dict[str, str], workbook) -> None:
        """Parse Rooms sheet (Excel only)."""
        sheet_key = normalize_sheet_name("Rooms")
        sheet_name = sheet_names[sheet_key]
        sheet = workbook[sheet_name]

        rows = self._sheet_to_dicts(sheet, sheet_name)
        if not rows:
            self.errors.append(f"ERROR [{sheet_name}]: Sheet is empty")
            return

        for row_num, row_dict in enumerate(rows, start=2):
            self._add_room_from_dict(row_dict, source=sheet_name, row_num=row_num)

    def _add_student_group_from_dict(self, group_dict: Dict[str, Any], source: str = "source", row_num: int = 0) -> None:
        """Add a student group from a dictionary (used by all format parsers)."""
        name = group_dict.get("name", "").strip() if isinstance(group_dict.get("name"), str) else ""

        if not name:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'name']: Required field is empty"
            )
            return

        if name in self.student_group_names:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + f", Column 'name']: Duplicate student group name '{name}'"
            )
            return

        size = coerce_int(group_dict.get("size"))
        if size is None or size <= 0:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'size']: "
                f"Expected positive integer, got {group_dict.get('size')}"
            )
            return

        level = (group_dict.get("level", "batch")).strip().lower() if isinstance(group_dict.get("level"), str) else "batch"
        if level not in VALID_LEVELS:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'level']: "
                f"Invalid level '{level}'. Must be one of: {', '.join(VALID_LEVELS)}"
            )
            return

        self.student_group_names.add(name)
        self.student_groups.append({
            "name": name,
            "size": size,
            "level": level,
        })

    def _parse_student_groups(self, sheet_names: Dict[str, str], workbook) -> None:
        """Parse Student Groups sheet (Excel only)."""
        sheet_key = normalize_sheet_name("Student Groups")
        sheet_name = sheet_names[sheet_key]
        sheet = workbook[sheet_name]

        rows = self._sheet_to_dicts(sheet, sheet_name)
        if not rows:
            self.errors.append(f"ERROR [{sheet_name}]: Sheet is empty")
            return

        for row_num, row_dict in enumerate(rows, start=2):
            self._add_student_group_from_dict(row_dict, source=sheet_name, row_num=row_num)

    def _add_course_from_dict(self, course_dict: Dict[str, Any], source: str = "source", row_num: int = 0) -> None:
        """Add a course from a dictionary (used by all format parsers)."""
        course_code = course_dict.get("course_code", "").strip() if isinstance(course_dict.get("course_code"), str) else ""
        session_type = course_dict.get("session_type", "").strip().lower() if isinstance(course_dict.get("session_type"), str) else ""
        professor = course_dict.get("professor", "").strip() if isinstance(course_dict.get("professor"), str) else ""
        room = course_dict.get("room", "").strip() if isinstance(course_dict.get("room"), str) else ""
        student_group = course_dict.get("student_group", "").strip() if isinstance(course_dict.get("student_group"), str) else ""

        required_fields = [
            ("course_code", course_code),
            ("session_type", session_type),
            ("professor", professor),
            ("room", room),
            ("student_group", student_group),
        ]

        for field_name, field_value in required_fields:
            if not field_value:
                self.errors.append(
                    f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + f", Column '{field_name}']: Required field is empty"
                )

        if session_type and session_type not in SESSION_TYPES:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + f", Column 'session_type']: "
                f"Invalid session type '{session_type}'. Must be one of: {', '.join(SESSION_TYPES)}"
            )

        if any(not v for _, v in required_fields) or session_type not in SESSION_TYPES:
            return

        course_tuple = (course_code, session_type, professor, student_group)
        if course_tuple in {(c["course_code"], c["session_type"], c["professor"], c["student_group"]) for c in self.courses}:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + f"]: "
                f"Duplicate course instance (course_code={course_code}, session_type={session_type}, professor={professor}, student_group={student_group})"
            )
            return

        course_name = course_dict.get("course_name", "").strip() if isinstance(course_dict.get("course_name"), str) else ""
        course_name = course_name or course_code
        slots_required = coerce_int(course_dict.get("slots_required")) or 1
        slots_continuous = coerce_bool(course_dict.get("slots_continuous"))
        preference_bin = coerce_int(course_dict.get("preference_bin")) or 1
        total_credits = coerce_int(course_dict.get("total_credits")) or 3

        if slots_required <= 0:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'slots_required']: "
                f"Expected positive integer, got {course_dict.get('slots_required')}"
            )
            return

        if preference_bin not in VALID_BINS:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'preference_bin']: "
                f"Invalid bin '{preference_bin}'. Must be one of: {', '.join(map(str, VALID_BINS))}"
            )
            return

        if total_credits <= 0:
            self.errors.append(
                f"ERROR [{source}" + (f", Row {row_num}" if row_num else "") + ", Column 'total_credits']: "
                f"Expected positive integer, got {course_dict.get('total_credits')}"
            )
            return

        self.courses.append({
            "course_code": course_code,
            "course_name": course_name,
            "session_type": session_type,
            "professor": professor,
            "room": room,
            "student_group": student_group,
            "slots_required": slots_required,
            "slots_continuous": slots_continuous,
            "preference_bin": preference_bin,
            "total_credits": total_credits,
        })

    def _parse_courses(self, sheet_names: Dict[str, str], workbook) -> None:
        """Parse Courses sheet (Excel only)."""
        sheet_key = normalize_sheet_name("Courses")
        sheet_name = sheet_names[sheet_key]
        sheet = workbook[sheet_name]

        rows = self._sheet_to_dicts(sheet, sheet_name)
        if not rows:
            self.errors.append(f"ERROR [{sheet_name}]: Sheet is empty")
            return

        for row_num, row_dict in enumerate(rows, start=2):
            self._add_course_from_dict(row_dict, source=sheet_name, row_num=row_num)

    def _cross_validate(self) -> None:
        """Validate cross-sheet references (professor/room/student_group)."""
        sheet_name = "Courses"
        for row_num, course in enumerate(self.courses, start=2):
            prof = course["professor"]
            if prof not in self.professor_names:
                self.errors.append(
                    f"ERROR [{sheet_name}, Row {row_num}, Column 'professor']: "
                    f"Unknown professor '{prof}'"
                )

            room = course["room"]
            if room not in self.room_names:
                self.errors.append(
                    f"ERROR [{sheet_name}, Row {row_num}, Column 'room']: "
                    f"Unknown room '{room}'"
                )

            group = course["student_group"]
            if group not in self.student_group_names:
                self.errors.append(
                    f"ERROR [{sheet_name}, Row {row_num}, Column 'student_group']: "
                    f"Unknown student group '{group}'"
                )

    def _sheet_to_dicts(self, sheet: Worksheet, sheet_name: str) -> List[Dict[str, str]]:
        """Convert worksheet rows to list of dicts (with normalized headers)."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [normalize_header(h) for h in rows[0]]
        data = []
        for row_values in rows[1:]:
            if all(v is None for v in row_values):
                continue
            row_dict = {headers[i]: (row_values[i] or "") for i in range(len(headers))}
            data.append(row_dict)

        return data

    def _build_stats(self) -> Dict[str, Any]:
        """Build statistics dictionary."""
        return {
            "counts": {
                "professors": len(self.professors),
                "rooms": len(self.rooms),
                "student_groups": len(self.student_groups),
                "courses": len(self.courses),
            },
            "errors": self.errors,
            "warnings": self.warnings,
        }

    def _write_json(self) -> None:
        """Write parsed data to JSON file."""
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        output = {
            "professors": self.professors,
            "rooms": self.rooms,
            "student_groups": self.student_groups,
            "courses": self.courses,
            "_meta": {
                "source_file": self.input_path.name,
                "parsed_at": datetime.now().isoformat(),
                "counts": {
                    "professors": len(self.professors),
                    "rooms": len(self.rooms),
                    "student_groups": len(self.student_groups),
                    "courses": len(self.courses),
                },
                "warnings": self.warnings,
            },
        }

        with open(self.output_path, "w") as f:
            json.dump(output, f, indent=2)

    def _print_errors(self) -> None:
        """Print all collected errors and warnings to stdout."""
        for error in self.errors:
            print(error)
        for warning in self.warnings:
            print(warning)


# ============================================================================
# CLI Entry Point
# ============================================================================

def main():
    """Command-line interface."""
    arg_parser = argparse.ArgumentParser(
        description="Parse timetable data (.xlsx, .csv, .json) and produce JSON intermediate format."
    )
    arg_parser.add_argument(
        "--input",
        "-i",
        required=True,
        help="Path to input file (.xlsx, .csv, or .json)",
    )
    arg_parser.add_argument(
        "--output",
        "-o",
        default="schema/output/parsed_data.json",
        help="Path to output JSON file (default: schema/output/parsed_data.json)",
    )
    arg_parser.add_argument(
        "--strict",
        action="store_true",
        help="Treat warnings as errors and fail if any are found",
    )

    args = arg_parser.parse_args()

    parser_obj = TimetableDataParser(strict=args.strict)
    success, stats = parser_obj.parse(args.input, args.output)

    print(f"\nParsing {'successful' if success else 'failed'}.")
    print(f"  Professors: {stats['counts']['professors']}")
    print(f"  Rooms: {stats['counts']['rooms']}")
    print(f"  Student Groups: {stats['counts']['student_groups']}")
    print(f"  Courses: {stats['counts']['courses']}")

    if stats["errors"]:
        print(f"\n❌ {len(stats['errors'])} error(s) found")
    if stats["warnings"]:
        print(f"\n⚠️  {len(stats['warnings'])} warning(s) found")

    if success:
        print(f"\n✅ Output written to: {args.output}")

    return 0 if success else 1


if __name__ == "__main__":
    exit(main())
