#!/usr/bin/env python
"""
Generate sample test Excel file for parser testing.

Uses seed data from schema/db_setup.py to create a properly formatted
Excel file with 4 sheets (Professors, Rooms, Student Groups, Courses).
"""

from pathlib import Path
from openpyxl import Workbook

# Seed data (from schema/db_setup.py)
PROFESSORS = [
    "Dr. Rakesh",
    "Dr. Sanjukta",
    "Dr. Anagha Tobi",
    "Dr. Shabnam",
    "Dr. Murtaza",
    "Dr. Naga Deepthi",
    "Dr. Satyanarayana Akula",
    "Dr. Ravi Kishor",
    "Dr. Neha",
    "Dr. A. Pujari",
    "Dr. Pradeep",
    "Dr. Anil Annadi",
    "Dr. OmPrakash",
    "Dr. Vandna Gokhroo",
    "Dr. Ramakant",
    "Dr. Kumudham",
    "Dr. Rakhee",
    "Dr. Sherlin",
    "Dr. Balaji Prashanth",
    "Dr. Nidhi Gupta",
    "Dr. P. Chaitanya Akshara",
    "Dr. Bhanu Kiran",
    "Dr. Sayantan",
    "Dr. Sabeeha",
    "Dr. Aruna Kumar Chelluboyina",
    "Dr. Sanjeev",
    "Dr. Akankasha Singh",
    "Dr. Varun Kumar",
    "Dr. Jayaprakash",
    "Dr. Manish",
    "Dr. Bill Reynolds",
    "Dr. Mittika",
    "Dr. Runa",
    "Dr. Anil",
    "Dr. Visalakshi",
    "Dr. Yayati",
    "Dr. Veeraiah",
    "Dr. Prafulla",
    "Dr. Praveen",
    "Dr. Satyanarayan",
    "Dr. Manoj",
    "Dr. Ravi Babu",
    "Dr. Yashwanth",
    "Dr. Bharghava",
    "Dr. Subbarao",
    "Dr. Mahesh",
    "Dr. Neeraj",
    "Dr. Avinash",
    "Dr. S Bharatala",
    "Dr. S Porika",
    "Dr. Raju",
    "Dr. Ankita",
    "Dr. Santosh Thakur",
    "Dr. Palash",
    "Dr. Prasad",
    "Dr. nartakannai",
]

ROOMS = [
    ("ELT 1", False, 120),
    ("ELT 2", False, 120),
    ("ELT 3", False, 120),
    ("ELT 4", False, 120),
    ("ELT 5", False, 120),
    ("E-LT 1", False, 100),
    ("E-LT 2", False, 100),
    ("E-LT 3", False, 100),
    ("E-LT 4", False, 100),
    ("E-LT 5", False, 100),
    ("ECR 2", False, 80),
    ("ECR 3", False, 80),
    ("ECR 5", False, 80),
    ("ECR 7", False, 80),
    ("ECR 8", False, 80),
    ("ECR 9", False, 80),
    ("ECR 10", False, 80),
    ("ECR 11", False, 80),
    ("ECR 13", False, 80),
    ("ECR 15", False, 80),
    ("ETR 1", False, 70),
    ("ETR 2", False, 70),
    ("ETR 3", False, 70),
    ("ETR 4", False, 70),
    ("ETR 5", False, 70),
    ("IT2 block", True, 90),
    ("CS LAB 1", True, 70),
    ("CS LAB 2", True, 70),
    ("CS LAB 3", True, 70),
    ("Comp Lab 3", True, 65),
    ("Material Testing Lab", True, 40),
    ("Geology Lab", True, 30),
    ("Auditorium", False, 400),
    ("E-Lab", False, 150),
]

STUDENT_GROUPS = [
    ("CS1", 83),
    ("CS2", 83),
    ("CS3", 83),
    ("CS4", 83),
    ("AI1", 70),
    ("AI2", 70),
    ("AI3", 70),
    ("ECE", 30),
    ("ECM", 60),
    ("CB", 15),
    ("BT", 40),
    ("CE", 10),
    ("ME", 25),
    ("MT", 25),
    ("NT", 5),
]

COURSES = [
    # CS1 courses
    ("MA2103", "Mathematics III", "lecture", "Dr. Rakesh", "ELT 1", "CS1", 5, False, 1, 3),
    ("HS2102", "Humanities II", "lecture", "Dr. Anagha Tobi", "ELT 2", "CS1", 4, False, 1, 3),
    ("HS2103", "Humanities III", "tutorial", "Dr. Kumudham", "ECR 2", "CS1", 2, False, 2, 3),
    ("CS/AI2101", "Programming", "lecture", "Dr. A. Pujari", "ELT 3", "CS1", 3, False, 2, 3),
    ("CS/AI 2102", "Data Structures", "lecture", "Dr. Ravi Kishor", "ELT 4", "CS1", 4, False, 2, 4),
    ("CS/AI 2102", "Data Structures", "lab", "Dr. Shabnam", "CS LAB 1", "CS1", 2, True, 2, 0),
    ("MA2104", "Discrete Math", "lecture", "Dr. Rakhee", "ELT 5", "CS1", 3, False, 1, 3),
    ("MA2105", "Linear Algebra", "lecture", "Dr. Rakesh", "E-LT 1", "CS1", 2, False, 1, 3),
    ("EC 2102", "Digital Logic", "lecture", "Dr. Sayantan", "ELT 2", "CS1", 4, False, 2, 3),
    ("EC 2102", "Digital Logic", "tutorial", "Dr. Satyanarayana Akula", "ECR 3", "CS1", 2, False, 2, 0),
    ("MA2106", "Calculus", "lecture", "Dr. Pradeep", "ELT 1", "CS1", 3, False, 1, 3),
    ("PH2102", "Physics II", "lecture", "Dr. Murtaza", "E-LT 2", "CS1", 4, False, 3, 3),
    ("PH2102", "Physics II", "tutorial", "Dr. Vandna Gokhroo", "ECR 5", "CS1", 2, False, 3, 0),
    ("CS2102", "Database Systems", "lab", "Dr. Neha", "CS LAB 2", "CS1", 2, True, 2, 0),

    # AI1 courses
    ("CS/AI 2102", "Data Structures", "lecture", "Dr. Ravi Kishor", "ELT 3", "AI1", 4, False, 2, 4),
    ("CS/AI 2102", "Data Structures", "lab", "Dr. Shabnam", "CS LAB 3", "AI1", 2, True, 2, 0),
    ("MA2103", "Mathematics III", "lecture", "Dr. Rakesh", "ELT 4", "AI1", 5, False, 1, 3),
    ("MA2103", "Mathematics III", "tutorial", "Dr. Rakesh", "ECR 7", "AI1", 2, False, 1, 0),
    ("CS/AI2101", "Programming", "lecture", "Dr. A. Pujari", "ELT 5", "AI1", 3, False, 2, 3),
    ("PH2102", "Physics II", "lab", "Dr. Murtaza", "Material Testing Lab", "AI1", 2, True, 3, 0),
    ("PH2102", "Physics II", "lecture", "Dr. Murtaza", "E-LT 3", "AI1", 3, False, 3, 3),
    ("EC 2102", "Digital Logic", "tutorial", "Dr. Ankita", "ECR 8", "AI1", 2, False, 2, 0),
    ("AI 2102", "Machine Learning", "lab", "Dr. Neha", "CS LAB 1", "AI1", 2, True, 2, 0),
    ("HS2103", "Humanities III", "tutorial", "Dr. Sherlin", "ECR 9", "AI1", 2, False, 1, 3),

    # ECE courses
    ("EC 2101", "Analog Circuits", "lecture", "Dr. P. Chaitanya Akshara", "ELT 1", "ECE", 4, False, 1, 3),
    ("EC 2101", "Analog Circuits", "lab", "Dr. P. Chaitanya Akshara", "Comp Lab 3", "ECE", 2, True, 1, 0),
    ("CS2104", "Web Tech", "lecture", "Dr. OmPrakash", "ELT 2", "ECE", 4, False, 2, 3),
    ("CS2104", "Web Tech", "lab", "Dr. nartakannai", "CS LAB 2", "ECE", 2, True, 2, 0),
    ("CS2104", "Web Tech", "tutorial", "Dr. OmPrakash", "ECR 10", "ECE", 2, False, 2, 0),
    ("MA2103", "Mathematics III", "lecture", "Dr. Rakesh", "ELT 3", "ECE", 4, False, 1, 3),
    ("MA2103", "Mathematics III", "tutorial", "Dr. Rakesh", "ECR 11", "ECE", 2, False, 1, 0),
    ("EC2103", "Control Systems", "lecture", "Dr. Bhanu Kiran", "ELT 4", "ECE", 3, False, 2, 3),
    ("EC 2102", "Digital Logic", "lecture", "Dr. Sayantan", "ELT 5", "ECE", 4, False, 2, 3),
    ("PH2102", "Physics II", "lecture", "Dr. Murtaza", "E-LT 1", "ECE", 4, False, 3, 3),
    ("PH2102", "Physics II", "tutorial", "Dr. Anil Annadi", "ECR 13", "ECE", 2, False, 3, 0),
    ("BT 2111", "Biotechnology", "lecture", "Dr. Sabeeha", "E-LT 2", "ECE", 3, False, 1, 3),
    ("BT 2111", "Biotechnology", "tutorial", "Dr. Sabeeha", "ECR 15", "ECE", 2, False, 1, 0),
]

def create_test_excel(output_path: str = "sample_timetable.xlsx"):
    """Create test Excel file with sample data."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Professors
    prof_sheet = wb.create_sheet("Professors")
    prof_sheet.append(["name", "email"])
    for prof in PROFESSORS:
        prof_sheet.append([prof, None])  # Email will be auto-generated
    print(f"✓ Professors sheet: {len(PROFESSORS)} professors")

    # Sheet 2: Rooms
    room_sheet = wb.create_sheet("Rooms")
    room_sheet.append(["name", "is_lab", "capacity"])
    for name, is_lab, capacity in ROOMS:
        room_sheet.append([name, is_lab, capacity])
    print(f"✓ Rooms sheet: {len(ROOMS)} rooms")

    # Sheet 3: Student Groups
    group_sheet = wb.create_sheet("Student Groups")
    group_sheet.append(["name", "size", "level"])
    for name, size in STUDENT_GROUPS:
        group_sheet.append([name, size, "batch"])
    print(f"✓ Student Groups sheet: {len(STUDENT_GROUPS)} groups")

    # Sheet 4: Courses
    course_sheet = wb.create_sheet("Courses")
    course_sheet.append([
        "course_code", "course_name", "session_type", "professor", "room",
        "student_group", "slots_required", "slots_continuous", "preference_bin", "total_credits"
    ])
    for course in COURSES:
        course_sheet.append(list(course))
    print(f"✓ Courses sheet: {len(COURSES)} course instances")

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"\n✅ Test Excel file created: {output_path}")
    print(f"   Total: {len(PROFESSORS)} professors, {len(ROOMS)} rooms, " +
          f"{len(STUDENT_GROUPS)} groups, {len(COURSES)} course instances")

    return output_path

if __name__ == "__main__":
    create_test_excel()
